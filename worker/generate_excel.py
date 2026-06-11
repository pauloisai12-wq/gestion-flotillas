# Generador de reportes Excel mensuales

import os
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from generate_pdf import (
    get_summary,
    get_fuel_by_type,
    get_top_consumers,
    get_kml_ranking,
    get_docs_summary,
    get_expired_docs_list,
    get_maintenance_done,
    get_maintenance_pending,
    MONTH_NAMES
)

# Carpeta donde se guardan los reportes
REPORTS_DIR = "/app/storage/reports"

# === ESTILOS ===
# Colores
BLUE_FILL = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
LIGHT_BLUE_FILL = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")
GREEN_FILL = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
RED_FILL = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FEF9C3", end_color="FEF9C3", fill_type="solid")
GRAY_FILL = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")

# Fuentes
TITLE_FONT = Font(name="Arial", size=14, bold=True, color="1E40AF")
HEADER_FONT = Font(name="Arial", size=10, bold=True, color="FFFFFF")
LABEL_FONT = Font(name="Arial", size=10, bold=True, color="334155")
VALUE_FONT = Font(name="Arial", size=10, color="1E293B")
SMALL_FONT = Font(name="Arial", size=9, color="64748B")

# Bordes
THIN_BORDER = Border(
    left=Side(style="thin", color="CBD5E1"),
    right=Side(style="thin", color="CBD5E1"),
    top=Side(style="thin", color="CBD5E1"),
    bottom=Side(style="thin", color="CBD5E1")
)


def style_header_row(ws, row_num, col_count):
    """
    Aplica estilo de encabezado a una fila (fondo azul, texto blanco).
    """
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.font = HEADER_FONT
        cell.fill = BLUE_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER


def style_data_rows(ws, start_row, end_row, col_count):
    """
    Aplica estilo a las filas de datos (bordes, filas alternas).
    """
    for row in range(start_row, end_row + 1):
        for col in range(1, col_count + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = VALUE_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="center")
            # Filas alternas con fondo gris
            if (row - start_row) % 2 == 1:
                cell.fill = GRAY_FILL


def auto_width(ws, col_count, min_width=12, max_width=35):
    """
    Ajusta automáticamente el ancho de las columnas según el contenido.
    """
    for col in range(1, col_count + 1):
        letter = get_column_letter(col)
        max_len = min_width
        for row in ws.iter_rows(min_col=col, max_col=col, values_only=False):
            for cell in row:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)) + 2)
        ws.column_dimensions[letter].width = min(max_len, max_width)


def write_table(ws, headers, data, start_row):
    """
    Escribe una tabla completa con encabezados y datos.
    Retorna el número de la siguiente fila disponible.
    """
    col_count = len(headers)

    # Escribir encabezados
    for col, header in enumerate(headers, 1):
        ws.cell(row=start_row, column=col, value=header)
    style_header_row(ws, start_row, col_count)

    # Escribir datos
    if not data:
        ws.cell(row=start_row + 1, column=1, value="Sin datos para este periodo")
        ws.cell(row=start_row + 1, column=1).font = SMALL_FONT
        return start_row + 3

    for row_idx, row_data in enumerate(data):
        row_num = start_row + 1 + row_idx
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_num, column=col_idx, value=value)

    end_row = start_row + len(data)
    style_data_rows(ws, start_row + 1, end_row, col_count)
    auto_width(ws, col_count)

    return end_row + 2


def create_resumen_sheet(wb, summary, docs_summary, month, year):
    """
    Hoja 1: Resumen ejecutivo con métricas clave.
    """
    ws = wb.active
    ws.title = "Resumen"

    # Título
    ws.cell(row=1, column=1, value=f"Reporte Mensual — {MONTH_NAMES[month]} {year}")
    ws.cell(row=1, column=1).font = TITLE_FONT
    ws.merge_cells("A1:D1")

    ws.cell(row=2, column=1, value=f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}")
    ws.cell(row=2, column=1).font = SMALL_FONT

    # Métricas en dos columnas: Label | Valor
    metrics = [
        ("Vehículos totales", summary["total_vehicles"]),
        ("Vehículos operativos", summary["operative_vehicles"]),
        ("Vehículos bloqueados", summary["blocked_vehicles"]),
        ("", ""),
        ("Cargas del mes", summary["total_loads"]),
        ("Litros totales", round(summary["total_liters"], 1)),
        ("Gasto total", f"${summary['total_spent']:,.2f}"),
        ("Presupuesto mensual", f"${summary['budget_total']:,.2f}"),
        ("Rendimiento promedio km/l", round(summary["avg_kml"], 2)),
        ("", ""),
        ("Documentos vigentes", docs_summary["valid"]),
        ("Documentos por vencer", docs_summary["expiring"]),
        ("Documentos vencidos", docs_summary["expired"]),
    ]

    row = 4
    for label, value in metrics:
        if label == "":
            row += 1
            continue
        ws.cell(row=row, column=1, value=label).font = LABEL_FONT
        ws.cell(row=row, column=2, value=value).font = VALUE_FONT
        ws.cell(row=row, column=1).border = THIN_BORDER
        ws.cell(row=row, column=2).border = THIN_BORDER

        # Colorear celdas según valor
        if label == "Vehículos bloqueados" and isinstance(value, (int, float)):
            if value > 0:
                ws.cell(row=row, column=2).fill = RED_FILL
            else:
                ws.cell(row=row, column=2).fill = GREEN_FILL
        if label == "Documentos vencidos" and isinstance(value, (int, float)):
            if value > 0:
                ws.cell(row=row, column=2).fill = RED_FILL
        if label == "Documentos por vencer" and isinstance(value, (int, float)):
            if value > 0:
                ws.cell(row=row, column=2).fill = YELLOW_FILL

        row += 1

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 20


def create_fuel_by_type_sheet(wb, fuel_by_type):
    """
    Hoja 2: Gasto de combustible por tipo de vehículo.
    """
    ws = wb.create_sheet("Combustible por Tipo")

    headers = ["Tipo de Vehículo", "Cargas", "Litros", "Gasto ($)", "km/l Promedio", "km/l Esperado"]
    data = []
    for row in fuel_by_type:
        data.append([
            row["vehicle_type"],
            row["total_loads"],
            round(row["total_liters"], 1),
            round(row["total_spent"], 2),
            round(row["avg_kml"], 2),
            round(row["expected_kml"], 2)
        ])

    write_table(ws, headers, data, start_row=1)


def create_top_consumers_sheet(wb, top_consumers):
    """
    Hoja 3: Top 10 vehículos con mayor gasto.
    """
    ws = wb.create_sheet("Top Consumidores")

    headers = ["#Eco", "Placa", "Tipo", "Cargas", "Litros", "Gasto ($)", "km/l"]
    data = []
    for row in top_consumers:
        data.append([
            row["economic_number"],
            row["plate"],
            row["vehicle_type"],
            row["total_loads"],
            round(row["total_liters"], 1),
            round(row["total_spent"], 2),
            round(row["avg_kml"], 2)
        ])

    write_table(ws, headers, data, start_row=1)


def create_ranking_sheet(wb, best_kml, worst_kml):
    """
    Hoja 4: Ranking de rendimiento km/l (mejores y peores).
    """
    ws = wb.create_sheet("Rendimiento km-l")

    # Título: Mejores
    ws.cell(row=1, column=1, value="Top 10 — Mejor Rendimiento km/l")
    ws.cell(row=1, column=1).font = Font(name="Arial", size=12, bold=True, color="166534")

    headers = ["#", "#Eco", "Placa", "Tipo", "km/l Real", "km/l Esperado", "Desviación %"]
    data_best = []
    for i, row in enumerate(best_kml, 1):
        data_best.append([
            i,
            row["economic_number"],
            row["plate"],
            row["vehicle_type"],
            round(row["avg_kml"], 2),
            round(row["expected_kml"], 2),
            f"+{round(row['deviation'], 1)}%"
        ])

    next_row = write_table(ws, headers, data_best, start_row=3)

    # Título: Peores
    ws.cell(row=next_row, column=1, value="Top 10 — Peor Rendimiento km/l")
    ws.cell(row=next_row, column=1).font = Font(name="Arial", size=12, bold=True, color="991B1B")

    data_worst = []
    for i, row in enumerate(worst_kml, 1):
        data_worst.append([
            i,
            row["economic_number"],
            row["plate"],
            row["vehicle_type"],
            round(row["avg_kml"], 2),
            round(row["expected_kml"], 2),
            f"{round(row['deviation'], 1)}%"
        ])

    write_table(ws, headers, data_worst, start_row=next_row + 2)


def create_docs_sheet(wb, expired_docs_list):
    """
    Hoja 5: Documentos vencidos.
    """
    ws = wb.create_sheet("Documentos Vencidos")

    headers = ["#Eco", "Placa", "Tipo Documento", "Fecha Vencimiento", "Días Vencido"]
    data = []
    for row in expired_docs_list:
        data.append([
            row["economic_number"],
            row["plate"],
            row["doc_type"],
            row["expires_at"],
            row["days_overdue"]
        ])

    write_table(ws, headers, data, start_row=1)


def create_maintenance_sheet(wb, maintenance_done, maintenance_pending):
    """
    Hoja 6: Mantenimientos realizados y pendientes.
    """
    ws = wb.create_sheet("Mantenimientos")

    # Título: Realizados
    ws.cell(row=1, column=1, value="Mantenimientos Realizados en el Mes")
    ws.cell(row=1, column=1).font = Font(name="Arial", size=12, bold=True, color="1E40AF")

    headers_done = ["#Eco", "Placa", "Servicio", "Fecha", "Odómetro", "Costo ($)", "Proveedor"]
    data_done = []
    for row in maintenance_done:
        data_done.append([
            row["economic_number"],
            row["plate"],
            row["service_name"],
            row["performed_at"],
            row["odometer_km"],
            row["cost"],
            row.get("provider", "")
        ])

    next_row = write_table(ws, headers_done, data_done, start_row=3)

    # Título: Pendientes
    ws.cell(row=next_row, column=1, value="Mantenimientos Pendientes / Vencidos")
    ws.cell(row=next_row, column=1).font = Font(name="Arial", size=12, bold=True, color="991B1B")

    headers_pending = ["#Eco", "Placa", "Servicio", "Km Actual", "Km Próximo Servicio", "Estado"]
    data_pending = []
    for row in maintenance_pending:
        data_pending.append([
            row["economic_number"],
            row["plate"],
            row["service_name"],
            row["current_km"],
            row["next_service_km"],
            "VENCIDO" if row["status"] == "OVERDUE" else "PROXIMO"
        ])

    write_table(ws, headers_pending, data_pending, start_row=next_row + 2)


def generate_excel(month, year, requested_by="sistema"):
    """
    Función principal: genera el reporte Excel del mes indicado.

    Parámetros:
        month: Número del mes (1-12)
        year: Año (ej: 2026)
        requested_by: Email del usuario que solicitó el reporte

    Retorna:
        Ruta del archivo Excel generado
    """
    print(f"  [Excel] Recopilando datos para {month}/{year}...")

    # 1. Recopilar datos (reutiliza las funciones de generate_pdf.py)
    summary = get_summary(month, year)
    fuel_by_type = get_fuel_by_type(month, year)
    top_consumers = get_top_consumers(month, year)
    best_kml, worst_kml = get_kml_ranking(month, year)
    docs_summary = get_docs_summary()
    expired_docs_list = get_expired_docs_list()
    maintenance_done = get_maintenance_done(month, year)
    maintenance_pending = get_maintenance_pending()

    print("  [Excel] Datos recopilados. Generando hojas...")

    # 2. Crear workbook
    wb = Workbook()

    # 3. Crear cada hoja
    create_resumen_sheet(wb, summary, docs_summary, month, year)
    create_fuel_by_type_sheet(wb, fuel_by_type)
    create_top_consumers_sheet(wb, top_consumers)
    create_ranking_sheet(wb, best_kml, worst_kml)
    create_docs_sheet(wb, expired_docs_list)
    create_maintenance_sheet(wb, maintenance_done, maintenance_pending)

    # 4. Guardar archivo
    os.makedirs(REPORTS_DIR, exist_ok=True)
    filename = f"reporte_mensual_{year}_{str(month).zfill(2)}.xlsx"
    filepath = os.path.join(REPORTS_DIR, filename)

    print(f"  [Excel] Guardando: {filename}...")
    wb.save(filepath)

    print(f"  [Excel] Excel generado: {filepath}")
    return filepath


# --- Prueba directa ---
if __name__ == "__main__":
    print("=== PRUEBA DE GENERACION DE EXCEL ===")
    path = generate_excel(month=3, year=2026, requested_by="admin@flotillas.com")
    print(f"\nArchivo generado en: {path}")